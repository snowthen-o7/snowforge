"""Build the formatted Excel version of the Feed Audit Template."""
import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule

wb = openpyxl.Workbook()

# ── Colors ──
COLORS = {
    "header_bg": "1a1a2e",
    "header_fg": "e8e8e8",
    "critical_bg": "fce4e4",
    "critical_fg": "991b1b",
    "high_bg": "fff3cd",
    "high_fg": "856404",
    "medium_bg": "e2e3e5",
    "medium_fg": "383d41",
    "low_bg": "f8f9fa",
    "low_fg": "6c757d",
    "pass_bg": "d4edda",
    "fail_bg": "f8d7da",
    "partial_bg": "fff3cd",
    "na_bg": "e2e3e5",
    "border": "d0d0d0",
    "alt_row": "f9f9fb",
}

thin_border = Border(
    left=Side(style="thin", color=COLORS["border"]),
    right=Side(style="thin", color=COLORS["border"]),
    top=Side(style="thin", color=COLORS["border"]),
    bottom=Side(style="thin", color=COLORS["border"]),
)

# ── Read CSV data ──
with open("audit-template.csv", "r", encoding="utf-8") as f:
    reader = csv.DictReader(f)
    rows = list(reader)

# ── Sheet 1: Audit Checklist ──
ws = wb.active
ws.title = "Audit Checklist"
ws.sheet_properties.tabColor = "ef4444"

# Freeze header row
ws.freeze_panes = "A2"

# Column config: (header, width, wrap)
columns = [
    ("Category", 22, False),
    ("Check", 55, True),
    ("Severity", 12, False),
    ("Platform", 14, False),
    ("Status", 12, False),
    ("Notes", 35, True),
    ("Fix Instructions", 60, True),
]

# Header row
header_font = Font(name="Calibri", bold=True, color=COLORS["header_fg"], size=11)
header_fill = PatternFill(start_color=COLORS["header_bg"], end_color=COLORS["header_bg"], fill_type="solid")
for col_idx, (header, width, _) in enumerate(columns, 1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border
    ws.column_dimensions[get_column_letter(col_idx)].width = width

# Data rows
severity_fills = {
    "Critical": PatternFill(start_color=COLORS["critical_bg"], end_color=COLORS["critical_bg"], fill_type="solid"),
    "High": PatternFill(start_color=COLORS["high_bg"], end_color=COLORS["high_bg"], fill_type="solid"),
    "Medium": PatternFill(start_color=COLORS["medium_bg"], end_color=COLORS["medium_bg"], fill_type="solid"),
    "Low": PatternFill(start_color=COLORS["low_bg"], end_color=COLORS["low_bg"], fill_type="solid"),
}
severity_fonts = {
    "Critical": Font(name="Calibri", bold=True, color=COLORS["critical_fg"], size=10),
    "High": Font(name="Calibri", bold=True, color=COLORS["high_fg"], size=10),
    "Medium": Font(name="Calibri", color=COLORS["medium_fg"], size=10),
    "Low": Font(name="Calibri", color=COLORS["low_fg"], size=10),
}

data_font = Font(name="Calibri", size=10)
alt_fill = PatternFill(start_color=COLORS["alt_row"], end_color=COLORS["alt_row"], fill_type="solid")

for row_idx, row_data in enumerate(rows, 2):
    values = [
        row_data.get("Category", ""),
        row_data.get("Check", ""),
        row_data.get("Severity", ""),
        row_data.get("Platform", ""),
        "",  # Status - left blank for user
        row_data.get("Notes", "").strip('"'),
        row_data.get("Fix_Instructions", "").strip('"'),
    ]
    for col_idx, val in enumerate(values, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.font = data_font
        cell.border = thin_border
        _, _, wrap = columns[col_idx - 1]
        cell.alignment = Alignment(wrap_text=wrap, vertical="top")

    # Alternating row background
    if row_idx % 2 == 0:
        for col_idx in range(1, len(columns) + 1):
            ws.cell(row=row_idx, column=col_idx).fill = alt_fill

    # Severity cell styling
    sev = values[2].strip()
    sev_cell = ws.cell(row=row_idx, column=3)
    if sev in severity_fills:
        sev_cell.fill = severity_fills[sev]
        sev_cell.font = severity_fonts[sev]
        sev_cell.alignment = Alignment(horizontal="center", vertical="top")

# Status dropdown validation
status_dv = DataValidation(
    type="list",
    formula1='"PASS,FAIL,PARTIAL,N/A"',
    allow_blank=True,
    showErrorMessage=True,
    errorTitle="Invalid Status",
    error="Please select PASS, FAIL, PARTIAL, or N/A",
)
status_dv.sqref = f"E2:E{len(rows) + 1}"
ws.add_data_validation(status_dv)

# Conditional formatting for Status column
ws.conditional_formatting.add(
    f"E2:E{len(rows) + 1}",
    CellIsRule(operator="equal", formula=['"PASS"'],
              fill=PatternFill(start_color=COLORS["pass_bg"], end_color=COLORS["pass_bg"], fill_type="solid"),
              font=Font(color="155724", bold=True)),
)
ws.conditional_formatting.add(
    f"E2:E{len(rows) + 1}",
    CellIsRule(operator="equal", formula=['"FAIL"'],
              fill=PatternFill(start_color=COLORS["fail_bg"], end_color=COLORS["fail_bg"], fill_type="solid"),
              font=Font(color="721c24", bold=True)),
)
ws.conditional_formatting.add(
    f"E2:E{len(rows) + 1}",
    CellIsRule(operator="equal", formula=['"PARTIAL"'],
              fill=PatternFill(start_color=COLORS["partial_bg"], end_color=COLORS["partial_bg"], fill_type="solid"),
              font=Font(color="856404", bold=True)),
)
ws.conditional_formatting.add(
    f"E2:E{len(rows) + 1}",
    CellIsRule(operator="equal", formula=['"N/A"'],
              fill=PatternFill(start_color=COLORS["na_bg"], end_color=COLORS["na_bg"], fill_type="solid"),
              font=Font(color="6c757d")),
)

# Auto-filter
ws.auto_filter.ref = f"A1:G{len(rows) + 1}"

# ── Sheet 2: Summary Dashboard ──
ws2 = wb.create_sheet("Summary")
ws2.sheet_properties.tabColor = "3b82f6"

summary_headers = ["Metric", "Value"]
for col_idx, h in enumerate(summary_headers, 1):
    cell = ws2.cell(row=1, column=col_idx, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border

ws2.column_dimensions["A"].width = 30
ws2.column_dimensions["B"].width = 20

summary_data = [
    ("Store / Brand Name", ""),
    ("Audit Date", ""),
    ("Auditor", ""),
    ("Platform", ""),
    ("", ""),
    ("Total Checks", f"=COUNTA('Audit Checklist'!A2:A{len(rows)+1})"),
    ("Checks Completed", f"=COUNTA('Audit Checklist'!E2:E{len(rows)+1})"),
    ("Completion %", f'=IF(B7=0,0,B8/B7)'),
    ("", ""),
    ("PASS", f"=COUNTIF('Audit Checklist'!E2:E{len(rows)+1},\"PASS\")"),
    ("FAIL", f"=COUNTIF('Audit Checklist'!E2:E{len(rows)+1},\"FAIL\")"),
    ("PARTIAL", f"=COUNTIF('Audit Checklist'!E2:E{len(rows)+1},\"PARTIAL\")"),
    ("N/A", f"=COUNTIF('Audit Checklist'!E2:E{len(rows)+1},\"N/A\")"),
    ("", ""),
    ("Critical Failures", f"=COUNTIFS('Audit Checklist'!E2:E{len(rows)+1},\"FAIL\",'Audit Checklist'!C2:C{len(rows)+1},\"Critical\")"),
    ("High Failures", f"=COUNTIFS('Audit Checklist'!E2:E{len(rows)+1},\"FAIL\",'Audit Checklist'!C2:C{len(rows)+1},\"High\")"),
    ("Medium Failures", f"=COUNTIFS('Audit Checklist'!E2:E{len(rows)+1},\"FAIL\",'Audit Checklist'!C2:C{len(rows)+1},\"Medium\")"),
    ("Low Failures", f"=COUNTIFS('Audit Checklist'!E2:E{len(rows)+1},\"FAIL\",'Audit Checklist'!C2:C{len(rows)+1},\"Low\")"),
    ("", ""),
    ("Overall Health Score", f'=IF(B7=0,0,(B11*1+B13*0.5)/(B7-B14))'),
]

for row_idx, (metric, value) in enumerate(summary_data, 2):
    cell_a = ws2.cell(row=row_idx, column=1, value=metric)
    cell_b = ws2.cell(row=row_idx, column=2, value=value)
    cell_a.font = Font(name="Calibri", bold=bool(metric), size=10)
    cell_b.font = Font(name="Calibri", size=10)
    cell_a.border = thin_border
    cell_b.border = thin_border
    if metric in ("Store / Brand Name", "Audit Date", "Auditor", "Platform"):
        cell_b.fill = PatternFill(start_color="fffde7", end_color="fffde7", fill_type="solid")

# Format percentages
ws2.cell(row=9, column=2).number_format = "0%"
ws2.cell(row=21, column=2).number_format = "0%"

# Highlight critical failures
ws2.conditional_formatting.add(
    "B16",
    CellIsRule(operator="greaterThan", formula=["0"],
              fill=PatternFill(start_color=COLORS["fail_bg"], end_color=COLORS["fail_bg"], fill_type="solid"),
              font=Font(color="721c24", bold=True)),
)

# ── Sheet 3: Instructions ──
ws3 = wb.create_sheet("How to Use")
ws3.sheet_properties.tabColor = "22c55e"
ws3.column_dimensions["A"].width = 100

instructions = [
    "E-COMMERCE FEED AUDIT TEMPLATE — QUICK START",
    "",
    "1. Go to the Summary tab and fill in your store name, date, and platform.",
    "",
    "2. Go to the Audit Checklist tab. Use the filters on Row 1 to focus on:",
    "   • Your platform (Shopify, WooCommerce, Google, Meta, or All)",
    "   • Severity level (start with Critical, then High)",
    "",
    "3. For each check, set the Status dropdown:",
    "   • PASS — Your feed meets this requirement",
    "   • FAIL — Your feed has this issue (fix it using the Fix Instructions column)",
    "   • PARTIAL — Partially met, needs improvement",
    "   • N/A — Not applicable to your store or feed",
    "",
    "4. Add notes in the Notes column for anything you want to remember.",
    "",
    "5. When done, check the Summary tab for your overall health score.",
    "",
    "PRIORITY ORDER:",
    "   1. Fix all Critical FAILs first — these cause disapprovals and block revenue",
    "   2. Fix High FAILs next — these degrade performance and waste ad spend",
    "   3. Address Medium items — these are optimization opportunities",
    "   4. Low items are nice-to-haves for feed completeness",
    "",
    "RECOMMENDED SCHEDULE:",
    "   • Weekly: Check Critical items (disapprovals, pricing, availability)",
    "   • Monthly: Full audit pass through all checks",
    "   • Quarterly: Review against latest platform spec changes",
    "",
    "For the full guide with examples and platform-specific advice,",
    "see the PDF guide included with your purchase.",
]

for row_idx, line in enumerate(instructions, 1):
    cell = ws3.cell(row=row_idx, column=1, value=line)
    if row_idx == 1:
        cell.font = Font(name="Calibri", bold=True, size=14)
    elif line.startswith(("PRIORITY", "RECOMMENDED")):
        cell.font = Font(name="Calibri", bold=True, size=11)
    else:
        cell.font = Font(name="Calibri", size=10)
    cell.alignment = Alignment(wrap_text=True)

# Save
output_path = "E-Commerce-Feed-Audit-Template.xlsx"
wb.save(output_path)
print(f"Created {output_path}")
print(f"  Sheets: {wb.sheetnames}")
print(f"  Audit rows: {len(rows)}")
print(f"  Features: dropdowns, conditional formatting, auto-filter, summary formulas")
